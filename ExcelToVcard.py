# importing openpyxl module
import openpyxl
import os

safeList = []
currentWorkingDir = os.path.abspath(os.getcwd())
# Fill static config i.e. values in double quotes
lastNameSuffix = "Wagh"
inputPath = currentWorkingDir + "/sample.xlsx"
outputPath = currentWorkingDir + "/output.vcf"

# workbook object is created
wb_obj = openpyxl.load_workbook(inputPath)
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

# Loop will print all columns name. Used to remove unwanted columns
for i in range(1, max_col + 1):
	cell_obj = sheet_obj.cell(row = 1, column = i)
	#print(cell_obj.value)
	if(cell_obj.value == "First Name of Student" or cell_obj.value == "Mobile number"):
		safeList.append(i)

colCount = 	max_col
while(colCount != 0):
	if not (safeList.__contains__(colCount)):
		sheet_obj.delete_cols(colCount)
	colCount = colCount - 1
	
# Adding new column name
sheet_obj.cell(row=1, column=3).value = "Last Name"

# Filling the entire column with suffix
for i in range(2, max_row + 1):
	if(sheet_obj.cell(row=i, column=1).value != None and sheet_obj.cell(row=i, column=1).value != ""):
		sheet_obj.cell(row=i, column=3).value = lastNameSuffix + " " + str(i)

# print full excel  
# for i in range(2, max_row+1):
#     row = [cell.value for cell in sheet_obj[i]] # sheet[n] gives nth row (list of cells)
#     print(row) # list of cell values of this row

def make_vcard():
    
	temp = []
	for i in range(2, max_row+1):
		if(sheet_obj.cell(row=i, column=1).value != None and sheet_obj.cell(row=i, column=1).value != ""):
			row = [cell.value for cell in sheet_obj[i]]
			first_name = row[0]
			phone = row[1]
			last_name = row[2]
			temp.append('BEGIN:VCARD')
			temp.append('VERSION:3.0')
			temp.append(f'N:{last_name};{first_name};;;')
			temp.append(f'FN:{first_name} {last_name}')
			temp.append(f'TEL;TYPE=VOICE,CELL;VALUE=text:{phone}')
			temp.append('END:VCARD')
	return temp

	# static return
    # return [
    #     'BEGIN:VCARD',
    #     'VERSION:3.0',
    #     f'N:Dov;John;;;',
    #     f'FN:John Dov',
    #     f'TEL;TYPE=VOICE,CELL;VALUE=text:1234567890',
    #     'END:VCARD',
    # ]

def write_vcard(f, vcard):
    with open(f, 'w') as f:
        f.writelines([l + '\n' for l in vcard])

vcard = make_vcard()

write_vcard(outputPath, vcard)

#save excel
#wb_obj.save(outputPath)

# Vcard Version 2.1
# return [
#         'BEGIN:VCARD',
#         'VERSION:2.1',
#         f'N:{last_name};{first_name}',
#         f'FN:{first_name} {last_name}',
#         f'ORG:{company}',
#         f'TITLE:{title}',
#         f'EMAIL;PREF;INTERNET:{email}',
#         f'TEL;WORK;VOICE:{phone}',
#         f'ADR;WORK;PREF:;;{address_formatted}',
#         f'REV:1',
#         'END:VCARD'
#     ]